import csv
import shutil
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from random import Random
from typing import Literal

import typer
from loguru import logger
from openpyxl import load_workbook

from street_sign_project.utils import project_root

# Define Path constants
DEFAULT_RAW_DATA_DIR = project_root() / "data" / "raw"

DEFAULT_PREPROCESS_DATA_DIR = project_root() / "data" / "preprocessed"
DEFAULT_CONFIG_DIR = project_root() / "configs"
DEFAULT_MAPPING_PATH_XLSX = DEFAULT_CONFIG_DIR / "street_sign_class_mapping.xlsx"
DEFAULT_MAPPING_PATH_CSV = DEFAULT_CONFIG_DIR / "street_sign_class_mapping.csv"

# Define other constants
MAPPING_SHEET_NAME = "canonical_mapping"
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png"}
POOL_DIR_NAME = "_split_pool"
DEFAULT_SPLIT_SEED = 420
DEFAULT_TRAIN_RATIO = 0.70
DEFAULT_VALID_RATIO = 0.15
DEFAULT_TEST_RATIO = 0.15

# Define Typing constants:
DatasetName = Literal["germany", "italy"]
SplitName = Literal["train", "valid", "test"]
SPLIT_NAMES: tuple[SplitName, ...] = ("train", "valid", "test")

app = typer.Typer(no_args_is_help=True)


@dataclass(frozen=True)
class ClassMapping:
    """Class names and dataset-specific class ID mappings."""

    class_names: tuple[str, ...]
    germany_to_canonical: dict[int, int]
    italy_to_canonical: dict[int, int]

    @classmethod
    def from_csv(cls, mapping_path: Path = DEFAULT_MAPPING_PATH_CSV) -> "ClassMapping":
        """Class constructor. Load class names and dataset-specific mappings from a CSV file."""
        class_names: list[str] = []
        germany_to_canonical: dict[int, int] = {}
        italy_to_canonical: dict[int, int] = {}

        logger.info(f"Loading class mapping from {mapping_path}")
        with mapping_path.open("r", encoding="utf-8", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                canonical_id = int(row["canonical_id"].strip())
                canonical_name = row["canonical_name"].strip()

                while len(class_names) <= canonical_id:
                    class_names.append("")
                class_names[canonical_id] = canonical_name

                italy_id = row["italy_id"].strip()
                if italy_id:
                    italy_to_canonical[int(italy_id)] = canonical_id

                germany_id = row["germany_id"].strip()
                if germany_id and germany_id.isdecimal():
                    germany_to_canonical[int(germany_id)] = canonical_id

        logger.info(f"Loaded {len(class_names)} canonical classes")
        return cls(
            class_names=tuple(class_names),
            germany_to_canonical=germany_to_canonical,
            italy_to_canonical=italy_to_canonical,
        )

    def to_canonical(self, dataset_name: DatasetName, class_id: int) -> int:
        """Return the canonical ID for a dataset specific class-id"""
        if dataset_name == "germany":
            return self.germany_to_canonical[class_id]
        elif dataset_name == "italy":
            return self.italy_to_canonical[class_id]

        raise ValueError(f"Unknown dataset name: {dataset_name}")


@dataclass(frozen=True)
class DatasetItem:
    """One image and its optional YOLO label file."""

    image_path: Path
    label_path: Path | None
    class_counts: dict[int, int]


def _remap_label_file(
    file_path: Path,
    output_path: Path,
    class_mapping: ClassMapping,
    dataset_name: DatasetName,
) -> None:
    """Remap one YOLO label file to canonical class IDs.

    Args:
        file_path: Path to the source label file.
        output_path: Path where the remapped label file should be written.
        class_mapping: Mapping used to convert dataset-specific class IDs.
        dataset_name: Source dataset name.

    Raises:
        ValueError: If a label line is malformed.
        KeyError: If a class ID has no canonical mapping.
    """
    remapped_lines: list[str] = []

    with file_path.open("r", encoding="utf-8") as label_file:
        # One file can have multiple lines
        for line_number, line in enumerate(label_file, start=1):
            values = line.strip().split()
            if not values:
                continue
            if len(values) != 5:
                raise ValueError(f"Expected 5 YOLO values in {file_path}:{line_number}, got {len(values)}.")

            # Get class id and remap to canonical
            class_id = int(values[0])
            try:
                canonical_id = class_mapping.to_canonical(dataset_name, class_id)
            except KeyError as error:
                raise KeyError(f"No canonical mapping for {dataset_name} class ID {class_id}.") from error

            # Add remapped lines to a list of strings
            remapped_lines.append(" ".join([str(canonical_id), *values[1:]]))

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Remap a list of lines to one string
    output_text = "\n".join(remapped_lines)
    output_text += "\n"  # Add empty line to the end (convention)

    with output_path.open("w", encoding="utf-8") as output_file:
        output_file.write(output_text)


def _remap_label_folder(
    input_labels_dir: Path,
    output_labels_dir: Path,
    class_mapping: ClassMapping,
    dataset_name: DatasetName,
    file_name_prefix: str = "",
) -> None:
    """Remap all YOLO label files in one folder to canonical values."""
    logger.info(f"Remapping {dataset_name} labels from {input_labels_dir} to {output_labels_dir}")

    remapped_count = 0
    for label_path in input_labels_dir.glob("*.txt"):
        output_path = output_labels_dir / f"{file_name_prefix}{label_path.name}"
        _remap_label_file(
            file_path=label_path,
            output_path=output_path,
            class_mapping=class_mapping,
            dataset_name=dataset_name,
        )
        remapped_count += 1

    logger.info(f"Remapped {remapped_count} label files for {dataset_name}")


def _copy_image_folder(input_images_dir: Path, output_images_dir: Path, file_name_prefix: str = "") -> None:
    """Copy all image files in one folder to another folder."""
    logger.info(f"Copying images from {input_images_dir} to {output_images_dir}")
    output_images_dir.mkdir(parents=True, exist_ok=True)

    copied_count = 0
    for image_path in input_images_dir.iterdir():
        if not image_path.is_file():
            continue
        if image_path.suffix.lower() not in IMAGE_SUFFIXES:
            continue

        shutil.copy2(image_path, output_images_dir / f"{file_name_prefix}{image_path.name}")
        copied_count += 1

    logger.info(f"Copied {copied_count} images to {output_images_dir}")


def _validate_split_ratios(train_ratio: float, valid_ratio: float, test_ratio: float) -> dict[SplitName, float]:
    """Validate split ratios and return them keyed by split name."""
    ratios: dict[SplitName, float] = {"train": train_ratio, "valid": valid_ratio, "test": test_ratio}
    if any(ratio <= 0 for ratio in ratios.values()):
        raise ValueError("All split ratios must be greater than zero.")
    if abs(sum(ratios.values()) - 1.0) > 1e-6:
        raise ValueError("Split ratios must sum to 1.0.")
    return ratios


def _target_split_sizes(total_items: int, ratios: dict[SplitName, float]) -> dict[SplitName, int]:
    """Calculate exact split sizes from fractional ratios. To create Train,val,test split"""
    raw_sizes = {split_name: total_items * ratios[split_name] for split_name in SPLIT_NAMES}
    split_sizes: dict[SplitName, int] = {
        split_name: int(raw_sizes[split_name]) for split_name in SPLIT_NAMES
    }  # Abrunden über int() wichtig
    remaining_items = total_items - sum(split_sizes.values())
    remainder_order = sorted(
        SPLIT_NAMES,
        key=lambda split_name: raw_sizes[split_name] - split_sizes[split_name],
        reverse=True,
    )
    for split_name in remainder_order[:remaining_items]:
        split_sizes[split_name] += 1
    return split_sizes


def _read_label_class_counts(label_path: Path | None) -> Counter[int]:
    """Count YOLO class IDs in a label file."""
    class_counts: Counter[int] = Counter()
    if label_path is None:
        return class_counts

    with label_path.open("r", encoding="utf-8") as label_file:
        for line_number, line in enumerate(label_file, start=1):
            values = line.strip().split()
            if not values:
                continue
            if len(values) != 5:
                raise ValueError(f"Expected 5 YOLO values in {label_path}:{line_number}, got {len(values)}.")
            class_counts[int(values[0])] += 1

    return class_counts


def _collect_dataset_items(images_dir: Path, labels_dir: Path) -> list[DatasetItem]:
    """Collect image and label pairs from a YOLO dataset directory."""
    items: list[DatasetItem] = []
    missing_label_count = 0

    for image_path in sorted(images_dir.iterdir()):
        if not image_path.is_file() or image_path.suffix.lower() not in IMAGE_SUFFIXES:
            continue

        label_path = labels_dir / f"{image_path.stem}.txt"
        if not label_path.exists():
            label_path = None
            missing_label_count += 1

        items.append(
            DatasetItem(
                image_path=image_path,
                label_path=label_path,
                class_counts=dict(_read_label_class_counts(label_path)),
            )
        )

    if missing_label_count:
        logger.warning(f"{missing_label_count} images have no matching label file.")

    return items


def _rarity_score(item: DatasetItem, total_class_counts: Counter[int]) -> float:
    """Score an item by how rare its classes are."""
    return sum(count / total_class_counts[class_id] for class_id, count in item.class_counts.items())


def _ordered_items_for_split(
    items: list[DatasetItem], total_class_counts: Counter[int], seed: int
) -> list[DatasetItem]:
    """Order items so rare-class images are assigned first with deterministic tie-breaking."""
    rng = Random(seed)
    items_with_tiebreakers = [(item, rng.random()) for item in items]
    items_with_tiebreakers.sort(
        key=lambda item_with_tiebreaker: (
            _rarity_score(item_with_tiebreaker[0], total_class_counts),
            len(item_with_tiebreaker[0].class_counts),
            item_with_tiebreaker[1],
        ),
        reverse=True,
    )
    return [item for item, _ in items_with_tiebreakers]


def _assignment_score(
    split_name: SplitName,
    item: DatasetItem,
    target_split_sizes: dict[SplitName, int],
    current_split_sizes: dict[SplitName, int],
    target_class_counts: dict[SplitName, dict[int, float]],
    current_class_counts: dict[SplitName, Counter[int]],
) -> float:
    """Score how well an item fits the current split deficits."""
    split_size_target = target_split_sizes[split_name]
    split_size_need = split_size_target - current_split_sizes[split_name]
    score = split_size_need / max(split_size_target, 1)

    for class_id, item_class_count in item.class_counts.items():
        class_target = target_class_counts[split_name].get(class_id, 0.0)
        class_need = class_target - current_class_counts[split_name][class_id]
        score += item_class_count * class_need / max(class_target, 1.0)

    return score


def _assign_dataset_items(
    items: list[DatasetItem],
    ratios: dict[SplitName, float],
    seed: int,
) -> dict[SplitName, list[DatasetItem]]:
    """Assign dataset items to train, valid, and test splits."""
    target_split_sizes = _target_split_sizes(total_items=len(items), ratios=ratios)
    total_class_counts: Counter[int] = Counter()
    for item in items:
        total_class_counts.update(item.class_counts)

    target_class_counts: dict[SplitName, dict[int, float]] = {
        split_name: {class_id: class_count * ratios[split_name] for class_id, class_count in total_class_counts.items()}
        for split_name in SPLIT_NAMES
    }
    current_class_counts: dict[SplitName, Counter] = {split_name: Counter() for split_name in SPLIT_NAMES}
    assignments: dict[SplitName, list[DatasetItem]] = {split_name: [] for split_name in SPLIT_NAMES}

    for item in _ordered_items_for_split(items=items, total_class_counts=total_class_counts, seed=seed):
        candidate_splits: list[SplitName] = [
            split_name for split_name in SPLIT_NAMES if len(assignments[split_name]) < target_split_sizes[split_name]
        ]
        best_split = max(
            candidate_splits,
            key=lambda split_name: _assignment_score(
                split_name=split_name,
                item=item,
                target_split_sizes=target_split_sizes,
                current_split_sizes={split: len(split_items) for split, split_items in assignments.items()},
                target_class_counts=target_class_counts,
                current_class_counts=current_class_counts,
            ),
        )
        assignments[best_split].append(item)
        current_class_counts[best_split].update(item.class_counts)

    return assignments


def _clear_split_dirs(output_dir: Path) -> None:
    """Remove generated split directories before recreating them."""
    for split_name in SPLIT_NAMES:
        split_dir = output_dir / split_name
        if split_dir.exists():
            shutil.rmtree(split_dir)


def _copy_dataset_items(assignments: dict[SplitName, list[DatasetItem]], output_dir: Path) -> None:
    """Copy assigned image and label pairs into split directories."""
    for split_name, items in assignments.items():
        output_images_dir = output_dir / split_name / "images"
        output_labels_dir = output_dir / split_name / "labels"
        output_images_dir.mkdir(parents=True, exist_ok=True)
        output_labels_dir.mkdir(parents=True, exist_ok=True)

        for item in sorted(items, key=lambda dataset_item: dataset_item.image_path.name):
            shutil.copy2(item.image_path, output_images_dir / item.image_path.name)
            if item.label_path is not None:
                shutil.copy2(item.label_path, output_labels_dir / item.label_path.name)

        logger.info(f"Created {split_name} split with {len(items)} images")


def _split_yolo_dataset(
    input_dir: Path,
    output_dir: Path,
    train_ratio: float = DEFAULT_TRAIN_RATIO,
    valid_ratio: float = DEFAULT_VALID_RATIO,
    test_ratio: float = DEFAULT_TEST_RATIO,
    seed: int = DEFAULT_SPLIT_SEED,
) -> None:
    """Split a pooled YOLO dataset into train, valid, and test directories."""
    ratios = _validate_split_ratios(train_ratio=train_ratio, valid_ratio=valid_ratio, test_ratio=test_ratio)
    items = _collect_dataset_items(images_dir=input_dir / "images", labels_dir=input_dir / "labels")
    if not items:
        raise ValueError(f"No images found in {input_dir / 'images'}")

    logger.info(
        f"Splitting {len(items)} images into train={train_ratio:.0%}, valid={valid_ratio:.0%}, test={test_ratio:.0%}"
    )
    assignments = _assign_dataset_items(items=items, ratios=ratios, seed=seed)
    _clear_split_dirs(output_dir=output_dir)
    _copy_dataset_items(assignments=assignments, output_dir=output_dir)


@app.command()
def export_mapping_to_csv(
    sheet_path: Path = DEFAULT_MAPPING_PATH_XLSX,
    output_path: Path = DEFAULT_MAPPING_PATH_CSV,
) -> None:
    """Export a mapping worksheet from an Excel workbook to a CSV file.

    Args:
        sheet_path: Path to the Excel workbook containing the mapping sheet.
        output_path: Path where the exported CSV file should be written.

    Raises:
        FileNotFoundError: If the Excel workbook does not exist.
        ValueError: If the canonical mapping worksheet is not present in the workbook.
    """
    logger.info(f"Exporting mapping sheet '{MAPPING_SHEET_NAME}' from {sheet_path} to {output_path}")

    if not sheet_path.exists():
        raise FileNotFoundError(f"Mapping workbook not found: {sheet_path}")

    # Load workbook
    workbook = load_workbook(sheet_path, data_only=True, read_only=True)
    try:
        if MAPPING_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Sheet '{MAPPING_SHEET_NAME}' not found in {sheet_path}.")

        output_path.parent.mkdir(parents=True, exist_ok=True)  # ensure outputpath exists
        worksheet = workbook[MAPPING_SHEET_NAME]

        # Write CSV file
        with output_path.open("w", encoding="utf-8", newline="") as csv_file:
            writer = csv.writer(csv_file)
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else value for value in row]
                if any(str(value).strip() for value in values):
                    writer.writerow(values)

        logger.info(f"Exported mapping CSV to {output_path}")
    finally:
        workbook.close()


@app.command()
def preprocess(
    raw_input_dir: Path = DEFAULT_RAW_DATA_DIR,
    output_dir: Path = DEFAULT_PREPROCESS_DATA_DIR,
    mapping_path: Path = DEFAULT_MAPPING_PATH_CSV,
    train_ratio: float = DEFAULT_TRAIN_RATIO,
    valid_ratio: float = DEFAULT_VALID_RATIO,
    test_ratio: float = DEFAULT_TEST_RATIO,
    seed: int = DEFAULT_SPLIT_SEED,
) -> None:
    """Remaps the labels and saves them to an output dir
    Assumes the raw structure, as it's currently present in both
    datasets.
    It first collects all data in one folder (pooled data)
    then does a deterministic 70/15/15 split.
    Uses a custom splitting logic, to ensure that each class is present in each split.
    Normal stratified methods do not work, as we can have multiple classes within one image.
    """
    logger.info(f"Preprocessing data from {raw_input_dir} into {output_dir}")

    # Initialize class mapping
    class_mapping = ClassMapping.from_csv(mapping_path)

    # Initialize pool dir and clear if it exsits
    pool_dir = output_dir / POOL_DIR_NAME
    if pool_dir.exists():
        shutil.rmtree(pool_dir)

    # Remap German Dataset into pooled directory
    for split in ["Train", "Test"]:
        input_split_dir = raw_input_dir / "GTSDB_Train_and_Test" / split
        file_name_prefix = f"germany_{split.lower()}_"
        _remap_label_folder(
            input_labels_dir=input_split_dir / "labels",
            output_labels_dir=pool_dir / "labels",
            class_mapping=class_mapping,
            dataset_name="germany",
            file_name_prefix=file_name_prefix,
        )
        _copy_image_folder(
            input_images_dir=input_split_dir / "images",
            output_images_dir=pool_dir / "images",
            file_name_prefix=file_name_prefix,
        )

    # Remap Italien Dataset into pooled directory
    for split in ["train", "test", "valid"]:
        input_split_dir = raw_input_dir / "StreetSignSet" / split
        file_name_prefix = f"italy_{split}_"
        _remap_label_folder(
            input_labels_dir=input_split_dir / "labels",
            output_labels_dir=pool_dir / "labels",
            class_mapping=class_mapping,
            dataset_name="italy",
            file_name_prefix=file_name_prefix,
        )
        _copy_image_folder(
            input_images_dir=input_split_dir / "images",
            output_images_dir=pool_dir / "images",
            file_name_prefix=file_name_prefix,
        )

    # Split data into train test val while preserving even rare classes in each split
    _split_yolo_dataset(
        input_dir=pool_dir,
        output_dir=output_dir,
        train_ratio=train_ratio,
        valid_ratio=valid_ratio,
        test_ratio=test_ratio,
        seed=seed,
    )
    shutil.rmtree(pool_dir)


@app.command()
def create_yaml_dataset(
    output_path: Path | None = None,
    mapping_path: Path = DEFAULT_MAPPING_PATH_CSV,
    dataset_path: str = "data/preprocessed",
    train_path: str = "train/images",
    valid_path: str = "valid/images",
    test_path: str = "test/images",
) -> None:
    """Create the YOLO dataset YAML file."""
    if output_path is None:
        output_path = project_root() / "data" / "dataset.yaml"

    mapping = ClassMapping.from_csv(mapping_path)
    ids_names_yaml = "\n".join(f"  {i}: {name}" for i, name in enumerate(mapping.class_names, start=0))
    yaml_content = (
        f"path: {dataset_path}\n"
        f"train: {train_path}\n"
        f"val: {valid_path}\n"
        f"test: {test_path}\n"
        f"\n"
        f"names:\n"
        f"{ids_names_yaml}\n"
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as file:
        file.write(yaml_content)


if __name__ == "__main__":
    app()
